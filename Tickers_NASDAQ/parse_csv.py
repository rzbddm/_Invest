import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from datetime import datetime
import openpyxl

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Hide the main tkinter window
Tk().withdraw()

# Ask for the CSV file
file_path = askopenfilename(title="Select CSV file", filetypes=[("CSV Files", "nasdaq_screener*.csv")], initialdir=r"C:\Users\a_h_J\OneDrive\Documents\_Invest\Tickers_NASDAQ")

if file_path:
    # Read the CSV file
    df = pd.read_csv(file_path)

    # Drop rows where 'Industry' is blank
    df = df[df['Industry'].notna()]

    # Drop specified columns
    columns_to_drop = ['Last Sale', 'Net Change', '% Change', 'Market Cap', 'Country', 'IPO Year', 'Volume']
    df.drop(columns=columns_to_drop, inplace=True)

    # Save the file with the new name format
    now = datetime.now().strftime("%Y-%m-%d_%H%M")
    output_file_path = f"NASDAQ_{now}.xlsx"

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='NASDAQ')

        # Adding other sheets with specified names
        sheet_names = ['Industrials', 'Real Estate', 'Finance', 'Health Care', 'Consumer Discretionary', 'Technology', 'Basic Materials', 'Consumer Staples', 'Energy', 'Miscellaneous', 'Utilities', 'Telecommunications']
        for sheet_name in sheet_names:
            count = df[df['Sector'] == sheet_name].shape[0]
            temp_df = pd.DataFrame({sheet_name: [sheet_name], 'Count': [count]})
            temp_df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Load the workbook and autofit columns
    workbook = openpyxl.load_workbook(output_file_path)
    ws_nasdaq = workbook['NASDAQ']
    ws_nasdaq.title = 'NASDAQ'
    autofit_columns(ws_nasdaq)

    for sheet_name in sheet_names:
        ws = workbook[sheet_name]
        ws['A1'] = sheet_name
        count = df[df['Sector'] == sheet_name].shape[0]
        ws['B1'] = count

        # Copy the data to the sheet starting from row 10
        sector_data = df[df['Sector'] == sheet_name]
        for idx, row in sector_data.iterrows():
            ws.append(row.values.tolist())

        # Delete row 2 in all sheets except NASDAQ
        ws.delete_rows(2)

        autofit_columns(ws)

    workbook.save(output_file_path)
    print(f"File saved as: {output_file_path}")

else:
    print("No file selected.")
